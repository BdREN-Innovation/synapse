#!/usr/bin/env python3
"""Verify that required spreadsheet cells contain evaluated formulas."""

import argparse
import json
import math
import sys
from pathlib import Path

from openpyxl import load_workbook

ERROR_VALUES = {"#VALUE!", "#DIV/0!", "#REF!", "#NAME?", "#NULL!", "#NUM!", "#N/A"}


def parse_cell(value):
    if "!" not in value:
        raise ValueError(f"Cell must use Sheet!A1 notation: {value}")
    sheet, coordinate = value.rsplit("!", 1)
    if not sheet or not coordinate:
        raise ValueError(f"Invalid cell reference: {value}")
    return sheet, coordinate


def parse_expected(value):
    cell, expected = value.split("=", 1)
    sheet, coordinate = parse_cell(cell)
    try:
        expected_value = json.loads(expected)
    except json.JSONDecodeError:
        expected_value = expected
    return sheet, coordinate, expected_value


def values_match(actual, expected):
    if isinstance(actual, (int, float)) and isinstance(expected, (int, float)):
        return math.isclose(actual, expected, rel_tol=1e-9, abs_tol=1e-9)
    return actual == expected


def main():
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("workbook")
    parser.add_argument("--formula", action="append", default=[], help="Required formula cell: Sheet!A1")
    parser.add_argument("--expect", action="append", default=[], help="Expected cached value: Sheet!A1=number or Sheet!A1=\"text\"")
    args = parser.parse_args()

    path = Path(args.workbook)
    if not path.is_file():
        raise ValueError(f"Workbook does not exist: {path}")

    formulas = load_workbook(path, data_only=False, read_only=True)
    values = load_workbook(path, data_only=True, read_only=True)
    failures = []
    checked = []
    try:
        for reference in args.formula:
            sheet, coordinate = parse_cell(reference)
            if sheet not in formulas.sheetnames:
                failures.append(f"missing sheet: {sheet}")
                continue
            formula = formulas[sheet][coordinate].value
            cached = values[sheet][coordinate].value
            checked.append(reference)
            if not (isinstance(formula, str) and formula.startswith("=")):
                failures.append(f"{reference} does not contain a formula")
            if cached is None:
                failures.append(f"{reference} has no cached calculated value")
            elif isinstance(cached, str) and any(error in cached for error in ERROR_VALUES):
                failures.append(f"{reference} contains Excel error {cached}")

        for item in args.expect:
            sheet, coordinate, expected = parse_expected(item)
            reference = f"{sheet}!{coordinate}"
            if sheet not in values.sheetnames:
                failures.append(f"missing sheet: {sheet}")
                continue
            actual = values[sheet][coordinate].value
            checked.append(reference)
            if actual is None:
                failures.append(f"{reference} has no cached value")
            elif not values_match(actual, expected):
                failures.append(f"{reference} expected {expected!r}, got {actual!r}")
    finally:
        formulas.close()
        values.close()

    result = {"status": "success" if not failures else "failed", "checked": checked, "failures": failures}
    print(json.dumps(result, indent=2))
    sys.exit(0 if not failures else 2)


if __name__ == "__main__":
    try:
        main()
    except Exception as error:
        print(json.dumps({"status": "failed", "failures": [str(error)]}, indent=2))
        sys.exit(1)
